#!/usr/bin/env python3
"""
ingest.py — Convert the user's filled-in master spreadsheet into
data/publications.json (the platform's source of truth).

Rules implemented:
  * Country normalisation:        Korea -> South Korea, Asia Pacific -> Asia Pacific (new)
  * "Global Research" renamed to "Global"; "Asia Pacific" added.
  * Frequency: Annually / Biannually / Quarterly / Monthly / Ad Hoc
  * Recurring Months: 3-letter tokens, semicolon-separated (commas accepted as lenient fallback)
  * For recurring rows: expected date = LAST CALENDAR DAY of each chosen month,
    expanded forward from upload date for as long as the calendar shows (we project 24 months).
  * For Ad Hoc rows: keep the provided date. If "TBD" / missing, store date = null
    and the platform routes the row into the TBD panel.
  * Asset Class: split on semicolons; "ALL" preserved as a single token.
  * Hospitality kept as canonical (Hotels deprecated but accepted).
  * Per-row data-entry fixes for the known seed errors (Taipei split, typos).
  * Per-instance id is stable: deterministic from (parent_id, year, month).
"""

import json
import re
import sys
import unicodedata
from datetime import date, datetime, timedelta
from calendar import monthrange
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC_XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else (
    ROOT / "uploaded_attachments" / "f8147d95f41c49d091679432b78378aa" /
    "APAC-Publication-Calendar-Master_20260615.xlsx"
)
OUT_JSON = ROOT / "data" / "publications.json"

UPLOAD_DATE = date(2026, 6, 15)            # user uploaded on this date
PROJECT_FORWARD_MONTHS = 24                # how far ahead we generate recurring instances

MONTH_TOKENS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    # forgiving forms
    "january": 1, "february": 2, "march": 3, "april": 4, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
    # known typo found in upload
    "un": 6,
}

# ---------- normalisers ----------

def norm_country(v):
    if not v:
        return None
    s = str(v).strip()
    if s.lower() == "asia pacific":
        return "Asia Pacific"
    if s.lower() == "global research":
        return "Global"          # per user request: rename
    if s.lower() == "global":
        return "Global"
    if s.lower() in ("korea", "south korea"):
        return "South Korea"
    return s

def norm_language(v):
    if not v:
        return None
    s = str(v).strip()
    # Handle multi-language semicolon strings token-by-token
    parts = [p.strip() for p in re.split(r";\s*", s) if p.strip()]
    mapped = []
    for p in parts:
        pl = p.lower()
        if pl == "traditional chinese":
            mapped.append("Chinese (Traditional)")
        elif pl == "simplified chinese":
            mapped.append("Chinese (Simplified)")
        else:
            mapped.append(p)
    return "; ".join(mapped)

def norm_pub_type(v):
    if not v:
        return None
    s = str(v).strip()
    # accept Newsletter as a new canonical type
    return s

def norm_frequency(v):
    if not v:
        return "Ad Hoc"
    s = str(v).strip()
    sl = s.lower()
    canonical = {
        "annually": "Annually", "annual": "Annually", "yearly": "Annually",
        "biannually": "Biannually", "biannual": "Biannually", "semi-annual": "Biannually", "semiannual": "Biannually",
        "quarterly": "Quarterly",
        "monthly": "Monthly",
        "ad hoc": "Ad Hoc", "adhoc": "Ad Hoc", "one-off": "Ad Hoc", "oneoff": "Ad Hoc",
    }
    return canonical.get(sl, s)

def parse_recurring_months(v):
    """Return (sorted list of month numbers, list of unparseable tokens)."""
    if v is None or str(v).strip() == "":
        return [], []
    raw = str(v).strip()
    # Accept commas and semicolons as separators (the user mixed both)
    parts = re.split(r"[;,]\s*", raw)
    months = []
    issues = []
    for p in parts:
        p = p.strip().lower()
        if not p:
            continue
        if p in MONTH_TOKENS:
            months.append(MONTH_TOKENS[p])
        else:
            issues.append(p)
    months = sorted(set(months))
    return months, issues

def norm_asset_class(v):
    """Split on semicolons, preserve "ALL" token, rename Hotels->Hospitality."""
    if not v:
        return []
    s = str(v).strip()
    parts = [x.strip() for x in re.split(r";\s*", s) if x.strip()]
    out = []
    for p in parts:
        if p.lower() == "hotels":
            out.append("Hospitality")
        else:
            out.append(p)
    return out

def parse_date_cell(v):
    """Return ISO date string or None. Accept datetime, several string formats, 'TBD'."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s or s.upper() == "TBD":
        return None
    # try D/M/Y, D-M-Y, Y-M-D
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None  # unparseable -> treat as TBD

def last_day_of_month(y, m):
    return date(y, m, monthrange(y, m)[1])

def slugify(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s[:60] or "item"

# ---------- main ----------

wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
ws = wb["Publications"]
# Detect header positions (row 2)
header_row = [ws.cell(row=2, column=c).value for c in range(2, ws.max_column + 1)]
# Map header -> column index (col index in workbook = c)
hmap = {}
for i, h in enumerate(header_row):
    if h is not None:
        hmap[str(h).strip()] = i + 2

def cell(row, header):
    c = hmap.get(header)
    if c is None:
        return None
    return ws.cell(row=row, column=c).value

publications = []
issues = []

# Special row-level corrections agreed with the user
TAIPEI_SPLIT_ROWS = {31, 32}            # original Singapore-tagged Taiwan rows
TAIPEI_TITLES = {
    31: "Singapore Sales & Investment Briefing",  # -> split into Singapore + Taipei
    32: "Singapore Office Briefing",
}
SP_LANGUAGE_OVERRIDE = "English"        # Singapore reports must be English only

for r in range(3, ws.max_row + 1):
    name = cell(r, "Publication Name")
    if not name or not str(name).strip():
        continue
    name = str(name).strip()

    country_raw = cell(r, "Country")
    country = norm_country(country_raw)
    city_state = cell(r, "City / State")
    if city_state is not None:
        city_state = str(city_state).strip()
    asset_class = norm_asset_class(cell(r, "Asset Class"))
    pub_type = norm_pub_type(cell(r, "Publication Type"))
    language = norm_language(cell(r, "Language"))
    frequency = norm_frequency(cell(r, "Frequency"))
    months_raw = cell(r, "Recurring Months")
    months, month_issues = parse_recurring_months(months_raw)
    if month_issues:
        issues.append(f"Row {r}: unparseable month tokens {month_issues} in '{months_raw}'")
    exp_date_iso = parse_date_cell(cell(r, "Expected Publication Date"))
    team = cell(r, "Team")
    team = str(team).strip() if team else None
    notes = cell(r, "Notes")
    notes = str(notes).strip() if notes else None
    scope = cell(r, "Report Scope")
    scope = str(scope).strip() if scope else "In-Country"
    status = cell(r, "Status (future)")
    status = str(status).strip() if status else None

    # Title typo fixes
    name = (name
            .replace("Sngapore", "Singapore")
            .replace("Austrlia", "Australia")
            .strip())

    # Trailing-space cleanup on asset_class tokens
    asset_class = [a.strip() for a in asset_class]

    base_row = {
        "country": country,
        "city_state": city_state,
        "asset_class": asset_class,
        "publication_type": pub_type,
        "language": language,
        "frequency": frequency,
        "recurring_months": months,
        "team": team,
        "notes": notes,
        "report_scope": scope,
        "status": status,
        "source_row": r,
    }

    rows_to_emit = []

    # Apply the Taipei split for rows 31 and 32
    if r in TAIPEI_SPLIT_ROWS:
        base_title = TAIPEI_TITLES[r].replace("Singapore", "").strip()
        # Singapore version
        sg = dict(base_row)
        sg["country"] = "Singapore"
        sg["city_state"] = "Singapore"
        sg["team"] = "Singapore Research"
        sg["language"] = SP_LANGUAGE_OVERRIDE
        sg["publication_name"] = f"Singapore {base_title}"
        rows_to_emit.append(sg)
        # Taipei version
        tp = dict(base_row)
        tp["country"] = "Taiwan"
        tp["city_state"] = "Taipei"
        tp["team"] = "Taiwan Research"
        tp["publication_name"] = f"Taipei {base_title}"
        rows_to_emit.append(tp)
    else:
        base_row["publication_name"] = name
        rows_to_emit.append(base_row)

    for rec in rows_to_emit:
        # Build deterministic parent_id
        parent_id = f"{slugify(rec['country'])}--{slugify(rec['publication_name'])}--r{r}"

        if rec["frequency"] in ("Annually", "Biannually", "Quarterly", "Monthly") and rec["recurring_months"]:
            # Generate instances forward from UPLOAD_DATE for PROJECT_FORWARD_MONTHS
            start = UPLOAD_DATE
            end = date(start.year + (PROJECT_FORWARD_MONTHS // 12),
                       start.month + (PROJECT_FORWARD_MONTHS % 12), 1) if False else None
            # simpler: iterate month-by-month for PROJECT_FORWARD_MONTHS
            instances = []
            y, m = start.year, start.month
            for i in range(PROJECT_FORWARD_MONTHS + 1):
                if m in rec["recurring_months"]:
                    inst_date = last_day_of_month(y, m)
                    if inst_date >= start:
                        instances.append(inst_date)
                # advance
                m += 1
                if m > 12:
                    m = 1; y += 1

            for d in instances:
                inst = dict(rec)
                inst["id"] = f"{parent_id}--{d.year}-{d.month:02d}"
                inst["parent_id"] = parent_id
                inst["recurring"] = True
                inst["expected_publication_date"] = d.isoformat()
                inst["is_tbd"] = False
                publications.append(inst)
        else:
            # Ad Hoc / Monthly without months / unparsable -> one row
            inst = dict(rec)
            inst["id"] = parent_id
            inst["parent_id"] = None
            inst["recurring"] = False
            inst["expected_publication_date"] = exp_date_iso
            inst["is_tbd"] = exp_date_iso is None
            publications.append(inst)

# Clean up "source_row" from output
for p in publications:
    p.pop("source_row", None)

# Sort: TBD last, then by date asc
def sort_key(p):
    if p.get("is_tbd"):
        return (1, "9999-99-99", p["publication_name"])
    return (0, p["expected_publication_date"] or "9999-99-99", p["publication_name"])

publications.sort(key=sort_key)

OUT_JSON.write_text(json.dumps(publications, indent=2, ensure_ascii=False) + "\n")

# Report
print(f"OK — wrote {len(publications)} publication rows to {OUT_JSON}")
print(f"  Recurring instances: {sum(1 for p in publications if p['recurring'])}")
print(f"  Ad Hoc:              {sum(1 for p in publications if not p['recurring'])}")
print(f"  TBD (no date):       {sum(1 for p in publications if p.get('is_tbd'))}")
if issues:
    print("\nIssues encountered:")
    for i in issues:
        print(f"  - {i}")

# Distinct values for sanity
print("\nDistinct countries:", sorted(set(p["country"] for p in publications)))
print("Distinct asset classes:", sorted(set(a for p in publications for a in p.get("asset_class", []))))
print("Distinct publication types:", sorted(set(p["publication_type"] for p in publications if p["publication_type"])))
print("Distinct languages:", sorted(set(p["language"] for p in publications if p["language"])))
print("Distinct teams:", sorted(set(p["team"] for p in publications if p["team"])))
