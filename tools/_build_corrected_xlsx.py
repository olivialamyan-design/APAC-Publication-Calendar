#!/usr/bin/env python3
"""Materialize the user-approved normalized JSON back into a master-format xlsx
so the canonical tools/build_publications.py parser can consume it unchanged.

We start from the RAW master workbook (to preserve the exact sheet layout,
headers, and every other cell verbatim) and overwrite ONLY the Publications
data cells (rows 3-36) with the normalized, pre-resolved values. This guarantees:
  * the 4 approved resolutions are reflected (rows 7, 10 dates; row 36 asset class),
  * Monthly rows 8 & 35 carry all 12 months so the existing expansion logic
    produces the same per-month instances it always has (the parser requires an
    explicit month list; it does not auto-expand a blank Monthly row),
  * no other value is altered.
"""
import json
import sys
from pathlib import Path

import openpyxl

RAW = Path(sys.argv[1])
NORM = Path(sys.argv[2])
OUT = Path(sys.argv[3])

ALL_MONTHS = "Jan; Feb; Mar; Apr; May; Jun; Jul; Aug; Sep; Oct; Nov; Dec"

norm = json.loads(NORM.read_text())
rows = {r["row"]: r for r in norm["rows"]}

wb = openpyxl.load_workbook(RAW)
ws = wb["Publications"]
cols = {ws.cell(row=2, column=c).value: c for c in range(1, ws.max_column + 1)}

# normalized field -> workbook header
FIELD_TO_HEADER = {
    "country": "Country",
    "cityState": "City / State",
    "publicationName": "Publication Name",
    "assetClass": "Asset Class",
    "publicationType": "Publication Type",
    "language": "Language",
    "frequency": "Frequency",
    "recurringMonths": "Recurring Months",
    "expectedPublicationDate": "Expected Publication Date",
    "startDate": "Start Date",
    "recurringTimingWindow": "Recurring Timing Window",
    "dateConfidence": "Date Confidence",
    "team": "Team",
    "leadAuthor": "Lead Author",
    "notes": "Notes",
    "reportScope": "Report Scope",
    "status": "Status (future)",
}


def set_cell(r, header, value):
    c = cols[header]
    ws.cell(row=r, column=c).value = value


for rno, rec in rows.items():
    recurring_months = rec["recurringMonths"]
    # Monthly rows had months cleared to null in the normalized JSON; the parser
    # needs an explicit list to expand them, so restore all 12 (== same result).
    if (rec.get("frequency") or "").strip().lower() == "monthly" and not recurring_months:
        recurring_months = ALL_MONTHS

    for field, header in FIELD_TO_HEADER.items():
        if header not in cols:
            continue
        if field == "recurringMonths":
            val = recurring_months
        else:
            val = rec.get(field)
        # write ISO date strings as plain strings; parser handles fromisoformat
        set_cell(rno, header, val if val is not None else None)

wb.save(OUT)
print(f"Wrote corrected workbook -> {OUT}")
