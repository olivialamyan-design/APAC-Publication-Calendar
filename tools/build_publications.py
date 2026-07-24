#!/usr/bin/env python3
"""Build data/publications.json from the APAC Publication Calendar master xlsx.

Design goals (per Olivia's working rules):
  * Map columns by HEADER NAME, not position — robust to column reorder.
  * Support `Lead Author`, `Start Date`, `Recurring Timing Window`, `Date Confidence`.
  * Expand recurring series with the ESTIMATED-DATE planning model (v4):
      - one record per occurrence,
      - the date of each occurrence is chosen by this precedence:
          1. Expected Publication Date  -> applies to the NEXT occurrence only,
             then later occurrences revert to estimated timing.
          2. Recurring Timing Window    -> placeholder date for the month:
               Early month -> 5th  (adjusted to a business day, same month)
               Mid-month   -> 15th (adjusted to a business day, same month)
               Late month  -> 25th (adjusted to a business day, same month)
               TBD / blank -> fall through to the fallback rule below.
          3. Fallback (no timing rule) -> LAST BUSINESS DAY of the month
             (weekdays only; public holidays are ignored for this logic).
      - projected forward over a fixed window (generation month .. +24 months),
      - occurrences before a series `Start Date` are skipped; if Start Date is
        blank, projection begins from the generation/upload date.
  * Date Confidence is owner-overridable; if blank it is auto-inferred:
        exact next-occurrence date -> Confirmed (that instance only)
        timing window present       -> Estimated
        neither                     -> TBD
  * Business day = weekday (Mon–Fri). Weekend dates move to the PREVIOUS
    business day, kept within the same month.
  * Output record shape is a superset of the existing publications.json
    (adds recurring_timing_window + date_confidence); all prior fields kept.

This script is intentionally dependency-light (openpyxl + stdlib) so it runs
the same locally and in CI.

Usage:
    python3 tools/build_publications.py \
        --xlsx APAC-Publication-Calendar-Master-5.xlsx \
        --out  data/publications.json
"""
import argparse
import calendar
import datetime
import json
import re
import sys
import unicodedata

import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

DATA_SHEET = "Publications"
HEADER_ROW = 2          # headers live on row 2
FIRST_DATA_ROW = 3      # data starts on row 3
FORWARD_MONTHS = 24     # projection window length (months ahead of gen month)

# Canonical field name  ->  list of acceptable header spellings (lower-cased).
# Header matching is case-insensitive and whitespace-insensitive.
HEADER_ALIASES = {
    "country":                   ["country"],
    "city_state":                ["city / state", "city/state", "city / state ", "city state"],
    "publication_name":          ["publication name"],
    "asset_class":               ["asset class"],
    "publication_type":          ["publication type"],
    "language":                  ["language"],
    "frequency":                 ["frequency"],
    "recurring_months":          ["recurring months"],
    "expected_publication_date": ["expected publication date"],
    "start_date":                ["start date"],
    "recurring_timing_window":   ["recurring timing window", "timing window"],
    "date_confidence":           ["date confidence"],
    "team":                      ["team"],
    "lead_author":               ["lead author"],
    "notes":                     ["notes"],
    "report_scope":              ["report scope"],
    "status":                    ["status (future)", "status"],
}

# Recurring Timing Window -> anchor day-of-month for the placeholder date.
# Each anchor is then adjusted to the previous business day within the month.
TIMING_WINDOW_ANCHOR = {
    "early month": 5,
    "mid-month":   15,
    "mid month":   15,
    "late month":  25,
}
TIMING_WINDOW_CANON = {
    "early month": "Early month",
    "mid-month":   "Mid-month",
    "mid month":   "Mid-month",
    "late month":  "Late month",
    "tbd":         "TBD",
}
DATE_CONFIDENCE_CANON = {
    "confirmed": "Confirmed",
    "estimated": "Estimated",
    "tbd":       "TBD",
}

MONTH_LOOKUP = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def slugify(text):
    """Lower-case, ASCII, hyphen-separated slug (matches existing ids)."""
    text = "" if text is None else str(text)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def norm_header(text):
    return re.sub(r"\s+", " ", str(text or "").strip().lower())


def last_day_of_month(year, month):
    return calendar.monthrange(year, month)[1]


def is_business_day(d):
    """Business day = weekday (Mon–Fri). Public holidays are intentionally
    ignored for recurring-date generation (per the v4 business rules)."""
    return d.weekday() < 5  # Mon=0 .. Fri=4


def adjust_to_business_day(year, month, day):
    """Return a date for (year, month, day) snapped to a business day, kept
    INSIDE the same month. If the anchor lands on a weekend, step BACKWARD to
    the previous business day. If stepping back would leave the month (e.g. the
    1st falls on a weekend), step FORWARD instead to the nearest business day
    within the month — i.e. the closest in-month business day."""
    day = max(1, min(day, last_day_of_month(year, month)))
    d = datetime.date(year, month, day)
    # Prefer the previous business day, but never cross out of the month.
    back = d
    while not is_business_day(back):
        prev = back - datetime.timedelta(days=1)
        if prev.month != month:
            break
        back = prev
    if is_business_day(back) and back.month == month:
        return back
    # Couldn't go back within the month -> step forward to the closest in-month
    # business day instead.
    fwd = d
    while not is_business_day(fwd) and fwd.month == month:
        fwd = fwd + datetime.timedelta(days=1)
    return fwd if (is_business_day(fwd) and fwd.month == month) else d


def last_business_day_of_month(year, month):
    """Fallback timing: last weekday of the month (was last CALENDAR day)."""
    d = datetime.date(year, month, last_day_of_month(year, month))
    while not is_business_day(d):
        d -= datetime.timedelta(days=1)
    return d


def placeholder_date_for_window(year, month, window_canon):
    """Map a canonical timing window to a business-day-adjusted date in the
    month, or None if the window does not provide a timing rule (TBD/blank)."""
    if not window_canon:
        return None
    key = window_canon.strip().lower()
    anchor = TIMING_WINDOW_ANCHOR.get(key)
    if anchor is None:
        return None  # 'TBD' (or unknown) -> no timing rule; caller uses fallback
    return adjust_to_business_day(year, month, anchor)


def add_months(year, month, n):
    """Return (year, month) n months after the given year/month."""
    idx = (year * 12 + (month - 1)) + n
    return idx // 12, idx % 12 + 1


def parse_date(value):
    """Accept datetime/date or 'YYYY-MM-DD' string -> date | None.

    The string 'TBD' (case-insensitive) returns None and flags TBD upstream.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    s = str(value).strip()
    if s.upper() == "TBD" or s == "":
        return None
    try:
        return datetime.date.fromisoformat(s[:10])
    except ValueError:
        return None


def is_tbd_value(value):
    return value is not None and str(value).strip().upper() == "TBD"


def parse_recurring_months(value, where=""):
    """Parse 'May; Sep' / 'Jan, Apr, Jul, Oct' / '2,5,8,11' -> sorted int list.

    Unrecognised tokens are reported to stderr (NOT silently corrected) so
    source-data typos surface to the operator rather than being papered over.
    """
    if value is None:
        return []
    tokens = re.split(r"[;,/]+|\s+and\s+", str(value))
    months = []
    for tok in tokens:
        tok = tok.strip().lower()
        if not tok:
            continue
        if tok.isdigit():
            m = int(tok)
            if 1 <= m <= 12:
                months.append(m)
            else:
                print(f"  ! {where}: month number out of range: {tok!r}", file=sys.stderr)
        elif tok in MONTH_LOOKUP:
            months.append(MONTH_LOOKUP[tok])
        elif tok[:3] in MONTH_LOOKUP:
            months.append(MONTH_LOOKUP[tok[:3]])
        else:
            print(f"  ! {where}: unrecognised month token {tok!r} in Recurring Months "
                  f"(check for typos, e.g. 'un' vs 'Jun')", file=sys.stderr)
    return sorted(set(months))


# Canonical asset-class spelling corrections. Applied after trimming; keeps the
# dynamically-derived filter bar clean when the master carries near-miss values.
# (2026-07-24) Owner-approved: 'Data Centre' -> 'Data Centres',
# 'Industrial' -> 'Industrial & Logistics'.
ASSET_CLASS_CANON = {
    "data centre":  "Data Centres",
    "data centres": "Data Centres",
    "industrial":   "Industrial & Logistics",
}


def parse_asset_class(value):
    """Asset class may hold multiple values separated by ; , / -> list[str].

    A lone 'TBD' placeholder is dropped (treated as blank) per owner direction.
    Trailing/leading whitespace is trimmed and near-miss spellings are snapped
    to the canonical taxonomy via ASSET_CLASS_CANON.
    """
    if value is None:
        return []
    parts = re.split(r"[;,/]+", str(value))
    out = []
    for p in parts:
        t = p.strip()
        if not t or t.upper() == "TBD":
            continue
        out.append(ASSET_CLASS_CANON.get(t.lower(), t))
    return out


def clean_text(value):
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def canon_timing_window(value, where=""):
    """Normalise a Recurring Timing Window cell to one of the four canonical
    values (Early month / Mid-month / Late month / TBD) or None if blank.
    Unknown non-blank values are reported and treated as None (fallback)."""
    s = clean_text(value)
    if not s:
        return None
    key = s.lower()
    canon = TIMING_WINDOW_CANON.get(key)
    if canon is None:
        print(f"  ! {where}: unrecognised Recurring Timing Window {s!r} "
              f"(use Early month / Mid-month / Late month / TBD) — using fallback",
              file=sys.stderr)
        return None
    return canon


def canon_date_confidence(value, where=""):
    """Normalise a Date Confidence cell to Confirmed / Estimated / TBD, or None
    if blank (None signals 'auto-infer'). Unknown values are reported."""
    s = clean_text(value)
    if not s:
        return None
    canon = DATE_CONFIDENCE_CANON.get(s.lower())
    if canon is None:
        print(f"  ! {where}: unrecognised Date Confidence {s!r} "
              f"(use Confirmed / Estimated / TBD) — will auto-infer",
              file=sys.stderr)
        return None
    return canon


# ---------------------------------------------------------------------------
# Core build
# ---------------------------------------------------------------------------

def build_header_map(ws):
    """Map canonical field -> column index (1-based) using header aliases."""
    found = {}
    for col in range(1, ws.max_column + 1):
        h = norm_header(ws.cell(row=HEADER_ROW, column=col).value)
        if not h:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if h in [norm_header(a) for a in aliases]:
                found[field] = col
                break
    missing = [f for f in ("country", "publication_name") if f not in found]
    if missing:
        raise SystemExit(f"Required column(s) not found by header name: {missing}")
    return found


def occurrences_in_window(recurring_months, gen_year, gen_month, start_date,
                          timing_window):
    """Compute every recurring occurrence inside
    [gen month .. gen month + FORWARD_MONTHS], honouring start_date.

    Each occurrence's ESTIMATED date is chosen by the timing window when one is
    given, else by the last-business-day fallback. Returns a list of dicts:
        {"year", "month", "date": date, "source": "timing"|"fallback"}
    sorted chronologically. The Expected-Publication-Date override for the NEXT
    occurrence is applied later by the caller.

    Eligibility uses the MONTH (not the placeholder day) so that an estimated
    date earlier in the month doesn't drop an otherwise-valid current-month
    occurrence.
    """
    end_year, end_month = add_months(gen_year, gen_month, FORWARD_MONTHS)
    win_end_month = datetime.date(end_year, end_month, 1)

    start_year_month = None
    if start_date is not None:
        start_year_month = datetime.date(start_date.year, start_date.month, 1)

    out = []
    y, m = gen_year, gen_month
    cur = datetime.date(y, m, 1)
    while cur <= win_end_month:
        if m in recurring_months:
            # Skip months strictly before the series Start Date's month.
            if start_year_month is None or cur >= start_year_month:
                est = placeholder_date_for_window(y, m, timing_window)
                if est is not None:
                    source = "timing"
                else:
                    est = last_business_day_of_month(y, m)
                    source = "fallback"
                # If Start Date is mid-month, only keep the start month when the
                # estimated date still falls on/after the Start Date itself.
                if start_date is not None and cur == start_year_month \
                        and est < start_date:
                    pass  # estimated slot already passed this month -> skip
                else:
                    out.append({"year": y, "month": m, "date": est, "source": source})
        y, m = add_months(y, m, 1)
        cur = datetime.date(y, m, 1)
    out.sort(key=lambda o: o["date"])
    return out


def build_records(ws, hmap, gen_date):
    gen_year, gen_month = gen_date.year, gen_date.month
    records = []

    def cell(row, field):
        col = hmap.get(field)
        return ws.cell(row=row, column=col).value if col else None

    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        name = clean_text(cell(row, "publication_name"))
        country = clean_text(cell(row, "country"))
        if not name and not country:
            continue  # blank row
        if not name:
            print(f"  ! row {row}: missing Publication Name — skipped", file=sys.stderr)
            continue

        country = country or ""
        frequency = clean_text(cell(row, "frequency")) or "Ad Hoc"
        recurring_months = parse_recurring_months(cell(row, "recurring_months"),
                                                   where=f"row {row} ({name})")
        start_date = parse_date(cell(row, "start_date"))
        timing_window = canon_timing_window(cell(row, "recurring_timing_window"),
                                            where=f"row {row} ({name})")
        owner_confidence = canon_date_confidence(cell(row, "date_confidence"),
                                                 where=f"row {row} ({name})")

        exp_raw = cell(row, "expected_publication_date")
        exp_date = parse_date(exp_raw)
        tbd = is_tbd_value(exp_raw) or (exp_date is None and exp_raw not in (None, ""))

        base = {
            "country": country,
            "city_state": clean_text(cell(row, "city_state")) or "",
            "asset_class": parse_asset_class(cell(row, "asset_class")),
            "publication_type": clean_text(cell(row, "publication_type")) or "",
            "language": clean_text(cell(row, "language")) or "English",
            "frequency": frequency,
            "recurring_months": recurring_months,
            "team": clean_text(cell(row, "team")) or "",
            "lead_author": clean_text(cell(row, "lead_author")) or "",
            "notes": clean_text(cell(row, "notes")),
            "report_scope": clean_text(cell(row, "report_scope")) or "",
            "status": clean_text(cell(row, "status")),
            "publication_name": name,
            # start_date is backend-only logic (not surfaced in the detail panel).
            "start_date": start_date.isoformat() if start_date else None,
            "recurring_timing_window": timing_window,  # canonical or None
        }

        # Stable id seed: r{spreadsheet row} keeps ids unique & deterministic.
        seed = f"{slugify(country)}--{slugify(name)}--r{row}"

        is_recurring = bool(recurring_months) and frequency.lower() not in ("ad hoc", "one-off", "")

        if is_recurring:
            occ = occurrences_in_window(recurring_months, gen_year, gen_month,
                                        start_date, timing_window)
            # ---- Exact next-occurrence override --------------------------------
            # Expected Publication Date, when given on a recurring row, replaces
            # the estimated date of the NEXT (earliest future) occurrence only.
            # Later occurrences keep their estimated timing.
            if exp_date is not None and occ:
                next_idx = None
                for i, o in enumerate(occ):
                    if o["date"] >= gen_date:
                        next_idx = i
                        break
                if next_idx is None:
                    next_idx = 0  # all within window are past -> use earliest
                occ[next_idx]["date"] = exp_date
                occ[next_idx]["source"] = "confirmed"
                occ.sort(key=lambda o: o["date"])

            for o in occ:
                d = o["date"]
                rec = dict(base)
                rec["id"] = f"{seed}--{d.strftime('%Y-%m')}"
                rec["parent_id"] = seed
                rec["recurring"] = True
                rec["expected_publication_date"] = d.isoformat()
                rec["is_tbd"] = False
                # ---- Date Confidence: owner override else auto-infer ----------
                if owner_confidence is not None:
                    rec["date_confidence"] = owner_confidence
                elif o["source"] == "confirmed":
                    rec["date_confidence"] = "Confirmed"
                else:  # 'timing' or 'fallback' -> estimated planning date
                    rec["date_confidence"] = "Estimated"
                records.append(rec)
            if not occ:
                print(f"  · row {row}: recurring series '{name}' produced 0 occurrences "
                      f"(check Start Date / Recurring Months)", file=sys.stderr)
        else:
            rec = dict(base)
            rec["id"] = seed
            rec["parent_id"] = None
            rec["recurring"] = False
            rec["expected_publication_date"] = exp_date.isoformat() if exp_date else None
            rec["is_tbd"] = bool(tbd)
            # Ad Hoc / one-off: confirmed if a date is set, else TBD (owner wins).
            if owner_confidence is not None:
                rec["date_confidence"] = owner_confidence
            elif exp_date is not None:
                rec["date_confidence"] = "Confirmed"
            else:
                rec["date_confidence"] = "TBD"
            records.append(rec)

    return records


def main():
    ap = argparse.ArgumentParser(description="Build publications.json from master xlsx")
    ap.add_argument("--xlsx", default="APAC-Publication-Calendar-Master-5.xlsx")
    ap.add_argument("--out", default="data/publications.json")
    ap.add_argument("--sheet", default=DATA_SHEET)
    ap.add_argument("--gen-date", default=None,
                    help="Override generation date (YYYY-MM-DD); defaults to today (UTC).")
    args = ap.parse_args()

    gen_date = (datetime.date.fromisoformat(args.gen_date)
                if args.gen_date else datetime.datetime.now(datetime.UTC).date())

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb[DATA_SHEET]
    hmap = build_header_map(ws)
    print(f"Header map: {hmap}")

    records = build_records(ws, hmap, gen_date)

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    n_rec = sum(1 for r in records if r["recurring"])
    n_one = len(records) - n_rec
    print(f"Wrote {len(records)} records ({n_one} one-off, {n_rec} recurring instances) "
          f"-> {args.out}  [gen month {gen_date:%Y-%m}, +{FORWARD_MONTHS}mo window]")


if __name__ == "__main__":
    main()
